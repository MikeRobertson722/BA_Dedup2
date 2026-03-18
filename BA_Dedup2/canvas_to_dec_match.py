"""
Source-to-DEC BA Matching Script
Compares Source BA records against DEC BA Master in database.
Scores BA match % (by SSN blocking) and Address match % separately.
Outputs results to import_merge_matches table.
"""
import re
import json
import os
import pandas as pd

import time
from openpyxl.styles import PatternFill, Font, Alignment

# Load .env file if present (no external dependency needed)
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _key, _val = _line.split('=', 1)
                os.environ[_key.strip()] = _val.strip()

from snowflake_conn import get_snowflake_connection
from config_loader import load_config, load_lookups

# Paths
SOURCE_FILE = 'input/SOURCE_BA_MASTER_.xlsx'

# Snowflake toggle
SNOWFLAKE_ENABLED = os.environ.get('SNOWFLAKE_ENABLED', 'false').lower() == 'true'

# API override: send ambiguous name matches (score 15-70) to Claude for evaluation
USE_API_OVERRIDE = True
API_MODEL = 'claude-sonnet-4-5-20250929'
API_SCORE_MIN = 15
API_SCORE_MAX = 70
API_BATCH_SIZE = 20

# Google Address Validation API: override ambiguous address scores (15-65)
USE_GOOGLE_ADDRESS_API = os.environ.get('GOOGLE_ADDRESS_API_ENABLED', 'false').lower() == 'true'
GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY', '')
GOOGLE_ADDR_SCORE_MIN = 15
GOOGLE_ADDR_SCORE_MAX = 65

# Classification thresholds (overridden by Snowflake config if available)
# ===== String matching functions =====

WORD_NUM = {
    "ONE": "1", "TWO": "2", "THREE": "3", "FOUR": "4", "FIVE": "5",
    "SIX": "6", "SEVEN": "7", "EIGHT": "8", "NINE": "9", "TEN": "10"
}


def safe_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def jaro(s1: str, s2: str) -> float:
    if s1 == s2:
        return 1.0
    len1, len2 = len(s1), len(s2)
    if len1 == 0 or len2 == 0:
        return 0.0
    match_dist = max(0, max(len1, len2) // 2 - 1)
    s1_matches = [False] * len1
    s2_matches = [False] * len2
    matches = 0
    transpositions = 0
    for i in range(len1):
        start = max(0, i - match_dist)
        end = min(i + match_dist + 1, len2)
        for j in range(start, end):
            if s2_matches[j] or s1[i] != s2[j]:
                continue
            s1_matches[i] = True
            s2_matches[j] = True
            matches += 1
            break
    if matches == 0:
        return 0.0
    k = 0
    for i in range(len1):
        if not s1_matches[i]:
            continue
        while not s2_matches[k]:
            k += 1
        if s1[i] != s2[k]:
            transpositions += 1
        k += 1
    return (matches / len1 + matches / len2 +
            (matches - transpositions / 2) / matches) / 3.0


def jaro_winkler(s1: str, s2: str, p: float = 0.1) -> float:
    j = jaro(s1, s2)
    prefix = 0
    for c1, c2 in zip(s1, s2):
        if c1 == c2:
            prefix += 1
        else:
            break
        if prefix == 4:
            break
    return j + prefix * p * (1 - j)


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return jaro_winkler(a, b)


def compact_alnum(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


# ===== Nickname / business-suffix / number-word lookups =====

NICKNAMES = {
    # Male
    'BILL': 'WILLIAM', 'BILLY': 'WILLIAM', 'WILL': 'WILLIAM', 'WILLY': 'WILLIAM',
    'WILLIE': 'WILLIAM', 'LIAM': 'WILLIAM',
    'BOB': 'ROBERT', 'BOBBY': 'ROBERT', 'ROB': 'ROBERT', 'ROBBIE': 'ROBERT',
    'CHUCK': 'CHARLES', 'CHARLIE': 'CHARLES', 'CHAS': 'CHARLES',
    'DICK': 'RICHARD', 'RICK': 'RICHARD', 'RICH': 'RICHARD', 'RICKY': 'RICHARD',
    'MIKE': 'MICHAEL', 'MIKEY': 'MICHAEL',
    'JIM': 'JAMES', 'JIMMY': 'JAMES', 'JIMMIE': 'JAMES', 'JAMIE': 'JAMES',
    'TOM': 'THOMAS', 'TOMMY': 'THOMAS',
    'JACK': 'JOHN', 'JOHNNY': 'JOHN', 'JON': 'JOHN',
    'JOE': 'JOSEPH', 'JOEY': 'JOSEPH',
    'ED': 'EDWARD', 'EDDIE': 'EDWARD', 'TED': 'EDWARD', 'TEDDY': 'EDWARD',
    'DAN': 'DANIEL', 'DANNY': 'DANIEL',
    'DAVE': 'DAVID', 'DAVY': 'DAVID',
    'STEVE': 'STEVEN', 'STEVIE': 'STEVEN', 'STEPHEN': 'STEVEN',
    'PAT': 'PATRICK',
    'TONY': 'ANTHONY',
    'LARRY': 'LAWRENCE', 'LARS': 'LAWRENCE',
    'HARRY': 'HAROLD', 'HAL': 'HAROLD',
    'JERRY': 'GERALD', 'GERRY': 'GERALD',
    'TERRY': 'TERRENCE',
    'MATT': 'MATTHEW',
    'ANDY': 'ANDREW', 'DREW': 'ANDREW',
    'FRED': 'FREDERICK', 'FREDDY': 'FREDERICK',
    'SAM': 'SAMUEL', 'SAMMY': 'SAMUEL',
    'BEN': 'BENJAMIN', 'BENNY': 'BENJAMIN',
    'KEN': 'KENNETH', 'KENNY': 'KENNETH',
    'RON': 'RONALD', 'RONNIE': 'RONALD',
    'DON': 'DONALD', 'DONNIE': 'DONALD',
    'PETE': 'PETER',
    'NICK': 'NICHOLAS',
    'WALT': 'WALTER', 'WALLY': 'WALTER',
    'HANK': 'HENRY',
    'RAY': 'RAYMOND',
    'PHIL': 'PHILIP', 'PHILLIP': 'PHILIP',
    'DOUG': 'DOUGLAS',
    'GREG': 'GREGORY',
    'JEFF': 'JEFFREY', 'GEOFF': 'JEFFREY',
    'GENE': 'EUGENE',
    'ART': 'ARTHUR',
    'VIC': 'VICTOR',
    'VINCE': 'VINCENT', 'VINNY': 'VINCENT',
    'STAN': 'STANLEY',
    'HERB': 'HERBERT',
    'LOU': 'LOUIS',
    'BERNIE': 'BERNARD',
    'MARTY': 'MARTIN',
    'NORM': 'NORMAN',
    'LENNY': 'LEONARD', 'LEO': 'LEONARD',
    'AL': 'ALBERT',
    'ALEX': 'ALEXANDER',
    'CHRIS': 'CHRISTOPHER',
    # Female
    'SANDY': 'SANDRA', 'SONDRA': 'SANDRA',
    'DOLORES': 'DELORES',
    'CATHY': 'CATHERINE', 'KATHY': 'CATHERINE', 'KATE': 'CATHERINE',
    'KATIE': 'CATHERINE', 'KATHERINE': 'CATHERINE', 'KATHRYN': 'CATHERINE',
    'BETTY': 'ELIZABETH', 'BETH': 'ELIZABETH', 'LIZ': 'ELIZABETH',
    'LIZZY': 'ELIZABETH', 'ELIZA': 'ELIZABETH',
    'PEGGY': 'MARGARET', 'PEG': 'MARGARET', 'MAGGIE': 'MARGARET',
    'MARGE': 'MARGARET', 'MARGIE': 'MARGARET',
    'SUE': 'SUSAN', 'SUZY': 'SUSAN', 'SUSIE': 'SUSAN',
    'BARB': 'BARBARA', 'BARBIE': 'BARBARA',
    'DEBBIE': 'DEBORAH', 'DEB': 'DEBORAH', 'DEBRA': 'DEBORAH',
    'JENNY': 'JENNIFER', 'JEN': 'JENNIFER',
    'SALLY': 'SARAH', 'SAL': 'SARAH',
    'TINA': 'CHRISTINA',
    'VICKY': 'VICTORIA', 'VIKKI': 'VICTORIA',
    'CONNIE': 'CONSTANCE',
    'DOTTY': 'DOROTHY', 'DOT': 'DOROTHY',
    'FRAN': 'FRANCES',
    'GINNY': 'VIRGINIA',
    'JUDY': 'JUDITH',
    'MILLIE': 'MILDRED',
    'NELL': 'HELEN', 'NELLIE': 'HELEN',
    'PATTY': 'PATRICIA', 'PATSY': 'PATRICIA', 'TRISH': 'PATRICIA',
}

BUSINESS_SUFFIXES = {
    'LLC', 'INC', 'CORP', 'CO', 'LTD', 'LP', 'LC', 'LLP',
    'PARTNERSHIP', 'PARTNE', 'PARTNERSHP', 'COMPANY',
}

# Business descriptors: interchangeable fluff when SSN already matches
BUSINESS_DESCRIPTORS = {
    'SERVICES', 'HOLDINGS', 'INVESTMENT', 'INVESTMENTS',
}

# Oil & gas industry terms: generic words that inflate scores between
# unrelated companies (e.g. OVINTIV EXPLORATION vs NEWFIELD EXPLORATION)
INDUSTRY_TERMS = {
    'EXPLORATION', 'DRILLING', 'PRODUCTION', 'PETROLEUM',
    'ENERGY', 'RESOURCES', 'OPERATING', 'PIPELINE',
    'MIDSTREAM', 'UPSTREAM', 'DOWNSTREAM',
    'MINERALS', 'ROYALTIES', 'PROPERTIES',
    'OIL', 'GAS', 'NATURAL',
}

NAME_SUFFIXES = {'JR', 'SR', 'II', 'III', 'IV', 'V'}

NUMBER_WORDS = {
    'ONE': '1', 'TWO': '2', 'THREE': '3', 'FOUR': '4', 'FIVE': '5',
    'SIX': '6', 'SEVEN': '7', 'EIGHT': '8', 'NINE': '9', 'TEN': '10',
}

# State abbreviation → full name (map abbrev TO full so both forms match;
# avoids conflict with CO in BUSINESS_SUFFIXES which strips before this runs)
STATE_ABBREVS = {
    'AL': 'ALABAMA', 'AK': 'ALASKA', 'AZ': 'ARIZONA', 'AR': 'ARKANSAS',
    'CA': 'CALIFORNIA', 'CT': 'CONNECTICUT',
    'DE': 'DELAWARE', 'FL': 'FLORIDA', 'GA': 'GEORGIA', 'HI': 'HAWAII',
    'ID': 'IDAHO', 'IL': 'ILLINOIS', 'IA': 'IOWA',
    'KS': 'KANSAS', 'KY': 'KENTUCKY', 'LA': 'LOUISIANA', 'ME': 'MAINE',
    'MD': 'MARYLAND', 'MA': 'MASSACHUSETTS', 'MI': 'MICHIGAN',
    'MN': 'MINNESOTA', 'MS': 'MISSISSIPPI', 'MO': 'MISSOURI',
    'MT': 'MONTANA', 'NE': 'NEBRASKA', 'NV': 'NEVADA',
    'NH': 'NEWHAMPSHIRE', 'NJ': 'NEWJERSEY', 'NM': 'NEWMEXICO',
    'NY': 'NEWYORK', 'NC': 'NORTHCAROLINA', 'ND': 'NORTHDAKOTA',
    'OH': 'OHIO', 'OK': 'OKLAHOMA', 'OR': 'OREGON', 'PA': 'PENNSYLVANIA',
    'RI': 'RHODEISLAND', 'SC': 'SOUTHCAROLINA', 'SD': 'SOUTHDAKOTA',
    'TN': 'TENNESSEE', 'TX': 'TEXAS', 'UT': 'UTAH', 'VT': 'VERMONT',
    'VA': 'VIRGINIA', 'WA': 'WASHINGTON', 'WV': 'WESTVIRGINIA',
    'WI': 'WISCONSIN', 'WY': 'WYOMING', 'DC': 'DISTRICTOFCOLUMBIA',
    # Skip IN (Indiana) and CO (Colorado) — too ambiguous as common words/suffixes
}


def canonicalize_token(token):
    """Normalize nicknames, number words, and state names to canonical form."""
    if token in NICKNAMES:
        return NICKNAMES[token]
    if token in NUMBER_WORDS:
        return NUMBER_WORDS[token]
    if token in STATE_ABBREVS:
        return STATE_ABBREVS[token]
    return token


def enhanced_token_overlap(tokens1, tokens2):
    """Token overlap counting exact, initial, and fuzzy (typo) matches."""
    if not tokens1 or not tokens2:
        return 0.0
    # Iterate over the smaller set, match against the larger
    if len(tokens1) > len(tokens2):
        small, big = list(tokens2), list(tokens1)
    else:
        small, big = list(tokens1), list(tokens2)

    matched = 0
    used = [False] * len(big)

    for t1 in small:
        found = False
        # 1) Exact match
        for j, t2 in enumerate(big):
            if not used[j] and t1 == t2:
                matched += 1
                used[j] = True
                found = True
                break
        if found:
            continue
        # 2) Initial match (single letter ↔ full name starting with that letter)
        if len(t1) == 1:
            for j, t2 in enumerate(big):
                if not used[j] and len(t2) > 1 and t2[0] == t1:
                    matched += 1
                    used[j] = True
                    found = True
                    break
        else:
            for j, t2 in enumerate(big):
                if not used[j] and len(t2) == 1 and t1[0] == t2:
                    matched += 1
                    used[j] = True
                    found = True
                    break
        if found:
            continue
        # 3) Fuzzy match for typos (both tokens > 4 chars, JW > 0.92)
        if len(t1) > 4:
            best_j, best_sim = -1, 0.0
            for j, t2 in enumerate(big):
                if not used[j] and len(t2) > 4:
                    sim = jaro_winkler(t1, t2)
                    if sim > best_sim:
                        best_sim = sim
                        best_j = j
            if best_sim > 0.88:
                matched += 1
                used[best_j] = True

    return matched / len(small)


# ===== Normalization functions =====

def normalize_name(name, trust: bool = False, detail: list | None = None) -> str:
    name = safe_str(name)
    if not name:
        if detail is not None:
            detail.append("EMPTY_INPUT")
        return ""
    if trust:
        if detail is not None:
            detail.append("TRUST_MODE: uppercase only")
        return re.sub(r"\s+", " ", name.upper()).strip()
    # Replace word-separating chars with space, strip all other specials
    before = name
    name = re.sub(r"[/&]", " ", name)
    name = re.sub(r"[^A-Za-z0-9\s]", "", name)
    if detail is not None and name != before:
        detail.append(f"[clean] {before.strip()}→{name.strip()}")
    suffix_map = {
        # Business name aliases (multi-word first)
        r"FEDERAL\s+EXPRESS": "FEDEX",
        r"FEDEX\s+OFFICE": "FEDEX",
        # Multi-word state names → single token (matches STATE_ABBREVS expansion)
        r"NEW\s+HAMPSHIRE": "NEWHAMPSHIRE",
        r"NEW\s+JERSEY": "NEWJERSEY",
        r"NEW\s+MEXICO": "NEWMEXICO",
        r"NEW\s+YORK": "NEWYORK",
        r"NORTH\s+CAROLINA": "NORTHCAROLINA",
        r"NORTH\s+DAKOTA": "NORTHDAKOTA",
        r"SOUTH\s+CAROLINA": "SOUTHCAROLINA",
        r"SOUTH\s+DAKOTA": "SOUTHDAKOTA",
        r"WEST\s+VIRGINIA": "WESTVIRGINIA",
        r"RHODE\s+ISLAND": "RHODEISLAND",
        r"DISTRICT\s+OF\s+COLUMBIA": "DISTRICTOFCOLUMBIA",
        # Trust abbreviation expansions (multi-word first)
        r"LIV\s+TR": "LIVING TRUST",
        "REVOC": "REVOCABLE",
        "REV": "REVOCABLE",
        "TR": "TRUST",
        # Name suffixes
        "JUNIOR": "JR", "SENIOR": "SR",
        "THIRD": "III", "FOURTH": "IV", "SECOND": "II",
        "CORPORATION": "CORP", "INCORPORATED": "INC",
        "COMPANY": "CO", "LIMITED": "LTD",
        r"L\s*L\s*C": "LLC", r"L\s*C": "LC", r"L\s*P": "LP",
    }
    for pat, repl in suffix_map.items():
        before = name
        name = re.sub(r"\b" + pat + r"\b", repl, name, flags=re.IGNORECASE)
        if detail is not None and name != before:
            matched = re.search(r"\b" + pat + r"\b", before, flags=re.IGNORECASE)
            detail.append(f"{matched.group() if matched else pat}→{repl}")
    # Split concatenated Roman numerals: PDIII → PD III
    before = name
    name = re.sub(r"([A-Za-z]{2,})(III|IV|II)", r"\1 \2", name)
    if detail is not None and name != before:
        detail.append(f"[roman_split] {before.strip()}→{name.strip()}")
    name = re.sub(r"\s+", " ", name.upper()).strip()
    # Collapse runs of consecutive single letters: "J V S LLC" → "JVS LLC"
    def _collapse_singles(m):
        return m.group(0).replace(" ", "")
    before = name
    name = re.sub(r"\b[A-Z](?:\s[A-Z]){1,}\b", _collapse_singles, name)
    if detail is not None and name != before:
        detail.append(f"[collapse] {before}→{name}")
    return name


# Date-like pattern: strips dates from addresses before comparison and number extraction
_MONTH_NAMES = (r'(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY'
                r'|JUNE?|JULY?|AUG(?:UST)?|SEP(?:T(?:EMBER)?)?'
                r'|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)')
_DATE_RE = re.compile(
    r'\b' + _MONTH_NAMES + r'[\s,]+\d{1,2}[\s,]+\d{2,4}\b'  # JUNE 14, 2012
    r'|\b\d{1,2}[\s,]+' + _MONTH_NAMES + r'[\s,]+\d{2,4}\b' # 14 JUNE 2012
    r'|\b' + _MONTH_NAMES + r'[\s,]+\d{4}\b'                 # MARCH 2012 (no day)
    r'|\b\d{1,2}[/\-\s]\d{1,2}[/\-\s]\d{2,4}\b'             # 1-23-14, MM/DD/YYYY
    r'|\b\d{8}\b'                                              # MMDDYYYY (8 digits)
    r'|\b(?:19|20)\d{2}\b'                                     # 4-digit year 19xx/20xx
    r'|\b\d{1,2}[/\-]\d{2,4}\b',                              # MM/YYYY or MM-YY
    re.IGNORECASE
)

# Patterns indicating a non-real / placeholder address → score 0
_UNKNOWN_ADDR_RE = re.compile(
    r'\b(?:'
    r'UNKNOWN\s+ADDR(?:ESS)?|ADDR(?:ESS)?\s+UNKNOWN'
    r'|NO\s+(?:VALID\s+)?ADDRESS'
    r'|NOT\s+AVAILABLE'
    r')\b'
    r'|^(?:NONE|N/?A|UNKNOWN)$'
)

# Placeholder/junk pattern for cities and other text fields
_UNKNOWN_TEXT_RE = re.compile(
    r'\b(?:UNKNOWN|NOT\s+AVAILABLE|N/?A)\b'
    r'|^(?:NONE|UNKNOWN)$'
)


def _needs_review_override(name, addr, city, state, zipcode,
                           dec_name=None, dec_addr=None, dec_city=None,
                           dec_state=None, dec_zip=None):
    """Return reason string if record has bad data requiring review, else None.

    Checks source fields (and optionally DEC fields) for: empty/null name or
    address, unknown city, empty city/state/zip, unknown address patterns.
    """
    # --- Source fields ---
    name_s = safe_str(name).strip()
    addr_s = safe_str(addr).strip()
    city_s = safe_str(city).strip()
    state_s = safe_str(state).strip()
    zip_s = safe_str(zipcode).strip()

    if not name_s:
        return 'EMPTY_NAME'
    if not addr_s:
        return 'EMPTY_ADDRESS'
    if not city_s:
        return 'EMPTY_CITY'
    if not state_s:
        return 'EMPTY_STATE'
    if not zip_s:
        return 'EMPTY_ZIP'
    if _UNKNOWN_ADDR_RE.search(addr_s.upper()):
        return 'UNKNOWN_ADDRESS'
    if _UNKNOWN_TEXT_RE.search(city_s.upper()):
        return 'UNKNOWN_CITY'
    if _UNKNOWN_TEXT_RE.search(name_s.upper()):
        return 'UNKNOWN_NAME'

    # --- DEC fields (only when a DEC match exists) ---
    if dec_name is not None:
        dn = safe_str(dec_name).strip()
        da = safe_str(dec_addr).strip()
        dc = safe_str(dec_city).strip()
        ds = safe_str(dec_state).strip()
        dz = safe_str(dec_zip).strip()

        if not dn:
            return 'DEC_EMPTY_NAME'
        if not da:
            return 'DEC_EMPTY_ADDRESS'
        if not dc:
            return 'DEC_EMPTY_CITY'
        if not ds:
            return 'DEC_EMPTY_STATE'
        if not dz:
            return 'DEC_EMPTY_ZIP'
        if _UNKNOWN_ADDR_RE.search(da.upper()):
            return 'DEC_UNKNOWN_ADDRESS'
        if _UNKNOWN_TEXT_RE.search(dc.upper()):
            return 'DEC_UNKNOWN_CITY'
        if _UNKNOWN_TEXT_RE.search(dn.upper()):
            return 'DEC_UNKNOWN_NAME'

    return None


def normalize_address(addr, trust: bool = False, detail: list | None = None) -> str:
    addr = safe_str(addr)
    if not addr:
        if detail is not None:
            detail.append("EMPTY_INPUT")
        return ""
    if trust:
        if detail is not None:
            detail.append("TRUST_MODE: uppercase only")
        return re.sub(r"\s+", " ", addr.upper()).strip()
    addr = addr.upper()

    # Strip "DATED <date>" patterns from address text (conservative — month names only,
    # NOT standalone 4-digit numbers which could be house numbers like 1950, 2000)
    before = addr
    addr = re.sub(r'\b(?:DATED|DTD)\s+', '', addr)
    addr = re.sub(
        r'\b' + _MONTH_NAMES + r'[\s,]+\d{1,2}[\s,]+\d{2,4}\b'
        r'|\b\d{1,2}[\s,]+' + _MONTH_NAMES + r'[\s,]+\d{2,4}\b'
        r'|\b' + _MONTH_NAMES + r'[\s,]+\d{4}\b',
        ' ', addr, flags=re.IGNORECASE)
    addr = re.sub(r'\s+', ' ', addr).strip()

    # Strip numeric dates (MM/DD/YYYY, MM-DD-YYYY, etc.) that survived month-name check
    addr = re.sub(r'\b\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\b', ' ', addr)
    addr = re.sub(r'\s+', ' ', addr).strip()
    if detail is not None and addr != before.upper().strip():
        detail.append(f"[date_strip] {before.strip()}→{addr}")

    # STRICT: Treat as PO Box ONLY if entire line is exactly "BOX <num/wordnum>"
    m = re.fullmatch(
        r"\s*BOX\s*#?\s*(\d+|ONE|TWO|THREE|FOUR|FIVE|SIX|SEVEN|EIGHT|NINE|TEN)\s*", addr)
    if m:
        tok = m.group(1).upper()
        tok = WORD_NUM.get(tok, tok)
        before = addr
        addr = f"PO BOX {tok}"
        if detail is not None:
            detail.append(f"[box_strict] {before}→{addr}")

    # Explicit PO patterns only
    before = addr
    addr = re.sub(r"\bPOST\s*OFFICE\s*BOX\b", "PO BOX", addr)
    addr = re.sub(r"\bP\.?\s*O\.?\s*BOX\b", "PO BOX", addr)
    addr = re.sub(r"\bP\.?\s*O\.?\s*B\b", "PO BOX", addr)
    addr = re.sub(r"\bPOB\b", "PO BOX", addr)
    addr = re.sub(r"\bPOP\s+BOX\b", "PO BOX", addr)
    addr = re.sub(r"\bPO\s+BOX\s*#", "PO BOX ", addr)
    if detail is not None and addr != before:
        detail.append(f"[pobox] {before}→{addr}")

    # Normalize word numbers in PO BOX addresses
    po_match = re.search(
        r"\bPO\s*BOX\s+(ONE|TWO|THREE|FOUR|FIVE|SIX|SEVEN|EIGHT|NINE|TEN)\b", addr)
    if po_match:
        word = po_match.group(1)
        before = addr
        addr = addr.replace(f"PO BOX {word}", f"PO BOX {WORD_NUM[word]}")
        if detail is not None:
            detail.append(f"PO BOX {word}→PO BOX {WORD_NUM[word]}")

    # Normalize COUNTY ROAD -> CR before street types (so ROAD doesn't get hit first)
    before = addr
    addr = re.sub(r"\bCOUNTY\s+ROAD\b", "CR", addr)
    addr = re.sub(r"\bCOUNTY\s+RD\b", "CR", addr)
    if detail is not None and addr != before:
        detail.append("COUNTY ROAD→CR")

    for full, abbr in [
        ("NORTHEAST", "NE"), ("NORTHWEST", "NW"),
        ("SOUTHEAST", "SE"), ("SOUTHWEST", "SW"),
        ("NORTH", "N"), ("SOUTH", "S"), ("EAST", "E"), ("WEST", "W"),
    ]:
        before = addr
        addr = re.sub(r"\b" + full + r"\b", abbr, addr)
        if detail is not None and addr != before:
            detail.append(f"{full}→{abbr}")

    for full, abbr in [
        ("STREET", "ST"), ("AVENUE", "AVE"), ("BOULEVARD", "BLVD"),
        ("DRIVE", "DR"), ("ROAD", "RD"), ("LANE", "LN"),
        ("COURT", "CT"), ("CIRCLE", "CIR"), ("PLACE", "PL"),
        ("HIGHWAY", "HWY"), ("PARKWAY", "PKWY"),
        ("TERRACE", "TER"), ("TRAIL", "TRL"),
    ]:
        before = addr
        addr = re.sub(r"\b" + full + r"\b", abbr, addr)
        if detail is not None and addr != before:
            detail.append(f"{full}→{abbr}")

    for full, abbr in [
        ("APARTMENT", "APT"), ("SUITE", "STE"),
        ("BUILDING", "BLDG"), ("FLOOR", "FL"), ("ROOM", "RM"),
        (r"#\s*(\d+)", r"APT \1"),
    ]:
        before = addr
        addr = re.sub(r"\b" + full + r"\b", abbr, addr)
        if detail is not None and addr != before:
            detail.append(f"{full}→{abbr}")

    before = addr
    addr = re.sub(r"[.,\-]", "", addr)
    if detail is not None and addr != before:
        detail.append(f"[punct] {before}→{addr}")

    # Strip leading non-address text (names, ATTN, C/O, titles, etc.)
    # Only for addresses that don't start with a digit and aren't PO Boxes.
    if not re.match(r'\s*\d', addr) and 'PO BOX' not in addr:
        _ADDR_START_TOKENS = {'HWY', 'CR', 'RR', 'RT', 'RTE', 'FM', 'SH', 'SR',
                              'CMR', 'PSC', 'HC', 'STAR',
                              'APT', 'STE', 'UNIT', 'BLDG', 'FL', 'RM'}
        m = re.search(r'\b(\d+)\s+[A-Z]', addr)
        if m:
            before_text = addr[:m.start()].strip()
            last_word = before_text.split()[-1] if before_text.split() else ''
            if last_word not in _ADDR_START_TOKENS:
                before = addr
                addr = addr[m.start():]
                if detail is not None:
                    detail.append(f"[strip_lead] {before}→{addr}")

    return re.sub(r"\s+", " ", addr).strip()


def normalize_city(val: str) -> str:
    s = safe_str(val).upper().strip()
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    # FT -> FORT (FT WORTH, FT SMITH, FT COLLINS, etc.)
    s = re.sub(r"\bFT\b", "FORT", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_zip(z: str) -> str:
    """Normalize zip code: strip to digits, left-pad to 5 if 4 digits."""
    digits = re.sub(r'[^0-9]', '', str(z)) if z else ''
    if len(digits) >= 4:
        return digits[:5].zfill(5)
    return digits


# ===== Parsing helpers =====

POBOX_CANON_RE = re.compile(r"\bPO\s*BOX\b", re.I)


def parse_house_number(addr_norm: str) -> str:
    m = re.match(r"^\s*(\d+)\b", addr_norm or "")
    return m.group(1) if m else ""


def parse_po_box_number(addr_norm: str) -> str:
    if not addr_norm or not POBOX_CANON_RE.search(addr_norm):
        return ""
    m = re.search(r"\bPO\s*BOX\s*([A-Z0-9]+)\b", addr_norm)
    if not m:
        return ""
    val = m.group(1).upper()
    return WORD_NUM.get(val, val) if val.isalpha() else val


def street_core_for_match(addr_norm: str) -> str:
    STREET_TYPE_TOKENS = {"ST", "AVE", "BLVD", "DR", "RD", "LN", "CT",
                          "CIR", "PL", "HWY", "PKWY", "TER", "TRL"}
    DIR_TOKENS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
    UNIT_TOKENS = {"APT", "STE", "UNIT", "BLDG", "FL", "RM"}
    s = (addr_norm or "").upper()
    s = re.sub(r"^\s*\d+\s+", "", s)
    s = re.sub(r"\b(APT|STE|UNIT|BLDG|FL|RM)\b\s*\w+", " ", s)
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    tokens = [t for t in s.split()
              if t not in STREET_TYPE_TOKENS
              and t not in DIR_TOKENS
              and t not in UNIT_TOKENS]
    return " ".join(tokens).strip()



def extract_addr_numbers(addr_norm: str) -> set:
    """Extract embedded numbers from address, ignoring date-like patterns."""
    cleaned = _DATE_RE.sub(' ', addr_norm or '')
    return set(re.findall(r'\d{2,}', cleaned))


# ===== Comparison functions =====

def name_compare(name1: str, name2: str, collect_detail: bool = False) -> dict:
    n1_detail = [] if collect_detail else None
    n2_detail = [] if collect_detail else None
    match_detail = [] if collect_detail else None

    n1 = normalize_name(name1, detail=n1_detail)
    n2 = normalize_name(name2, detail=n2_detail)
    if not n1 or not n2:
        result = {"name_score": 0.0, "name_match": False}
        if collect_detail:
            if match_detail is not None:
                match_detail.append("empty_input")
            result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
        return result
    if n1 == n2:
        result = {"name_score": 1.0, "name_match": True}
        if collect_detail:
            if match_detail is not None:
                match_detail.append("exact_post_norm")
            result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
        return result

    # Strip personal name suffixes (safe: SSN already matched)
    t1_raw = [t for t in n1.split() if t not in NAME_SUFFIXES]
    t2_raw = [t for t in n2.split() if t not in NAME_SUFFIXES]
    if not t1_raw or not t2_raw:
        t1_raw, t2_raw = n1.split(), n2.split()
    elif match_detail is not None:
        stripped1 = [t for t in n1.split() if t in NAME_SUFFIXES]
        stripped2 = [t for t in n2.split() if t in NAME_SUFFIXES]
        if stripped1 or stripped2:
            match_detail.append(f"strip_suffix: {', '.join(stripped1 + stripped2)}")

    # Acronym detection BEFORE business suffix stripping
    # e.g. CSC = CORP SERVICE CO (C-S-C)
    if len(t1_raw) == 1 and len(t1_raw[0]) > 1 and len(t2_raw) >= 2:
        acr = t1_raw[0]
        if len(acr) == len(t2_raw) and all(a == w[0] for a, w in zip(acr, t2_raw)):
            result = {"name_score": 0.95, "name_match": True}
            if match_detail is not None:
                match_detail.append(f"acronym: {acr}={' '.join(t2_raw)}→95")
                result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
            return result
    if len(t2_raw) == 1 and len(t2_raw[0]) > 1 and len(t1_raw) >= 2:
        acr = t2_raw[0]
        if len(acr) == len(t1_raw) and all(a == w[0] for a, w in zip(acr, t1_raw)):
            result = {"name_score": 0.95, "name_match": True}
            if match_detail is not None:
                match_detail.append(f"acronym: {acr}={' '.join(t1_raw)}→95")
                result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
            return result

    # Strip business suffixes/descriptors, remove AND, canonicalize nicknames/numbers
    STRIP = BUSINESS_SUFFIXES | BUSINESS_DESCRIPTORS | INDUSTRY_TERMS | {'AND', 'THE'}
    t1 = []
    for t in t1_raw:
        if t not in STRIP:
            ct = canonicalize_token(t)
            if match_detail is not None and ct != t:
                match_detail.append(f"canon: {t}→{ct}")
            t1.append(ct)
        elif match_detail is not None:
            match_detail.append(f"strip_biz: {t}")
    t2 = []
    for t in t2_raw:
        if t not in STRIP:
            ct = canonicalize_token(t)
            if match_detail is not None and ct != t:
                match_detail.append(f"canon: {t}→{ct}")
            t2.append(ct)
        elif match_detail is not None:
            match_detail.append(f"strip_biz: {t}")
    if not t1 or not t2:
        t1 = [canonicalize_token(t) for t in t1_raw]
        t2 = [canonicalize_token(t) for t in t2_raw]

    c1 = " ".join(t1)
    c2 = " ".join(t2)

    if c1 == c2:
        result = {"name_score": 1.0, "name_match": True}
        if match_detail is not None:
            match_detail.append("exact_post_strip")
            result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
        return result

    # Initial-match check (J SMITH vs JOHN SMITH)
    if len(t1) >= 2 and len(t2) >= 2 and t1[-1] == t2[-1]:
        if len(t1[0]) == 1 and len(t2[0]) > 1 and t1[0] == t2[0][0]:
            result = {"name_score": 0.95, "name_match": True}
            if match_detail is not None:
                match_detail.append(f"initial: {t1[0]}={t2[0]}→95")
                result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
            return result
        if len(t2[0]) == 1 and len(t1[0]) > 1 and t2[0] == t1[0][0]:
            result = {"name_score": 0.95, "name_match": True}
            if match_detail is not None:
                match_detail.append(f"initial: {t2[0]}={t1[0]}→95")
                result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
            return result

    # Character-level similarities (multiple strategies)
    direct_jw = jaro_winkler(c1, c2)
    sorted_jw = jaro_winkler(" ".join(sorted(t1)), " ".join(sorted(t2)))
    compact_jw = jaro_winkler(compact_alnum(c1), compact_alnum(c2))
    best_jw = max(direct_jw, sorted_jw, compact_jw)

    # Enhanced token overlap (counts initial matches and fuzzy/typo matches)
    if min(len(t1), len(t2)) > 1:
        overlap_coeff = enhanced_token_overlap(t1, t2)

        # Bookend bonus: if first AND last tokens match, middle name
        # differences shouldn't tank the score (SSN already matched)
        bookend = False
        if len(t1) >= 2 and len(t2) >= 2:
            first_match = (t1[0] == t2[0]
                           or (len(t1[0]) == 1 and len(t2[0]) > 1 and t1[0] == t2[0][0])
                           or (len(t2[0]) == 1 and len(t1[0]) > 1 and t2[0] == t1[0][0])
                           or (len(t1[0]) > 1 and len(t2[0]) > 1 and jaro_winkler(t1[0], t2[0]) >= 0.85))
            last_match = t1[-1] == t2[-1]
            if first_match and last_match:
                overlap_coeff = max(overlap_coeff, 0.9)
                bookend = True

        score = best_jw * (0.3 + 0.7 * overlap_coeff)
        if match_detail is not None:
            match_detail.append(f"jw={best_jw:.3f} (d={direct_jw:.3f}/s={sorted_jw:.3f}/c={compact_jw:.3f})")
            match_detail.append(f"overlap={overlap_coeff:.3f}")
            if bookend:
                match_detail.append("bookend_bonus")
    else:
        score = best_jw
        if match_detail is not None:
            match_detail.append(f"jw={best_jw:.3f} (d={direct_jw:.3f}/s={sorted_jw:.3f}/c={compact_jw:.3f})")

    # Compact similarity floor: handles compound words and concatenated initials
    if compact_jw >= 0.95:
        old_score = score
        score = max(score, compact_jw)
        if match_detail is not None and score != old_score:
            match_detail.append(f"compact_floor: {compact_jw:.3f}")

    if match_detail is not None:
        match_detail.append(f"score={score:.3f}")

    result = {"name_score": score, "name_match": score >= 0.85}
    if collect_detail:
        result.update(_build_name_detail(n1_detail, n2_detail, match_detail))
    return result


def _build_name_detail(n1_detail, n2_detail, match_detail):
    """Build the name detail strings from collected lists."""
    norm_parts = []
    if n1_detail:
        norm_parts.append(f"source: {'; '.join(n1_detail)}")
    if n2_detail:
        norm_parts.append(f"dec: {'; '.join(n2_detail)}")
    return {
        'name_norm_detail': ' | '.join(norm_parts) if norm_parts else '',
        'name_match_detail': '; '.join(match_detail) if match_detail else '',
    }


# Regex for extracting names from address fields (C/O, ATTN patterns)
_CO_RE = re.compile(
    r'\bC\s*/?\s*O\s+(.+?)(?:\s{2,}|\s+\d|\s+P\s*\.?\s*O\s|\s+BOX\s|$)',
    re.IGNORECASE)
_ATTN_RE = re.compile(
    r'\bATTN:?\s+(.+?)(?:\s{2,}|\s+\d|\s+P\s*\.?\s*O\s|\s+BOX\s|$)',
    re.IGNORECASE)
_ROLE_SUFFIX_RE = re.compile(
    r',?\s*(MANAGER|AGENT|DIRECTOR|TREASURER|SECRETARY|OFFICER|'
    r'PRESIDENT|VP|VICE PRESIDENT|ADMINISTRATOR|SUPERVISOR)\s*$',
    re.IGNORECASE)


def extract_names_from_address(addr: str) -> list:
    """Extract embedded person/entity names from C/O and ATTN patterns."""
    if not addr:
        return []
    text = safe_str(addr).upper()
    names = []
    for pattern in (_CO_RE, _ATTN_RE):
        m = pattern.search(text)
        if m:
            name = m.group(1).strip().rstrip(',')
            name = _ROLE_SUFFIX_RE.sub('', name).strip()
            if len(name) >= 3:
                names.append(name)
    return names


def address_compare(addr1, city1, zip1, addr2, city2, zip2,
                    collect_detail: bool = False) -> dict:
    a1_detail = [] if collect_detail else None
    a2_detail = [] if collect_detail else None
    match_detail = [] if collect_detail else None

    addr1_norm = normalize_address(addr1, detail=a1_detail)
    addr2_norm = normalize_address(addr2, detail=a2_detail)
    zip1 = normalize_zip(zip1)
    zip2 = normalize_zip(zip2)

    def _finalize(result):
        if collect_detail:
            result.update(_build_addr_detail(a1_detail, a2_detail, match_detail))
        return result

    # Placeholder / unknown address → score 0 (check before empty guards
    # so "source empty + DEC unknown" doesn't slip through as ONE_EMPTY)
    if ((addr1_norm and _UNKNOWN_ADDR_RE.search(addr1_norm)) or
            (addr2_norm and _UNKNOWN_ADDR_RE.search(addr2_norm))):
        if match_detail is not None:
            match_detail.append("UNKNOWN_ADDRESS")
        return _finalize({"same_address": False, "score": 0.0, "reason": "UNKNOWN_ADDRESS"})

    # Both empty
    if not addr1_norm and not addr2_norm:
        if match_detail is not None:
            match_detail.append("BOTH_EMPTY")
        return _finalize({"same_address": False, "score": 0.0, "reason": "BOTH_EMPTY"})
    if not addr1_norm or not addr2_norm:
        if match_detail is not None:
            match_detail.append("ONE_EMPTY")
        return _finalize({"same_address": False, "score": 0.0, "reason": "ONE_EMPTY"})

    is_pobox1 = bool(POBOX_CANON_RE.search(addr1_norm))
    is_pobox2 = bool(POBOX_CANON_RE.search(addr2_norm))

    if is_pobox1 or is_pobox2:
        if not (is_pobox1 and is_pobox2):
            if match_detail is not None:
                match_detail.append("pobox_vs_street")
            return _finalize({"same_address": False, "score": 0.0, "reason": "POBOX_VS_STREET"})
        box1 = parse_po_box_number(addr1_norm)
        box2 = parse_po_box_number(addr2_norm)
        if not box1 or not box2:
            if match_detail is not None:
                match_detail.append(f"pobox: box1={box1} box2={box2} MISSING_NUM")
            return _finalize({"same_address": False, "score": 0.0, "reason": "POBOX_MISSING_NUM"})
        if box1 != box2:
            if match_detail is not None:
                match_detail.append(f"pobox: {box1}!={box2}")
            return _finalize({"same_address": False, "score": 0.0,
                    "reason": f"POBOX_NUM_MISMATCH {box1}!={box2}"})

        city1_norm = normalize_city(city1)
        city2_norm = normalize_city(city2)
        if city1_norm and city2_norm:
            city_sim = max(similarity(city1_norm, city2_norm),
                          similarity(compact_alnum(city1_norm),
                                     compact_alnum(city2_norm)))
        else:
            city_sim = 0.90

        zip_bonus = 0.05 if (zip1 and zip2 and
                             zip1 == zip2) else 0.0
        score = min(1.0, 0.92 * city_sim + zip_bonus)
        same = score >= 0.90 and city_sim >= 0.85
        if match_detail is not None:
            match_detail.append(f"pobox: box={box1} MATCH; city_sim={city_sim:.2f}; zip_bonus={zip_bonus:.2f}; score={score:.3f}")
        return _finalize({"same_address": same, "score": score,
                "reason": f"POBOX_MATCH box={box1} city_sim={city_sim:.2f}"})

    # Street address comparison
    num1 = parse_house_number(addr1_norm)
    num2 = parse_house_number(addr2_norm)

    # If neither has a house number (military: CMR/PSC/UNIT, or other
    # non-standard formats), fall back to full normalized string comparison
    if not num1 and not num2:
        addr_sim = max(similarity(addr1_norm, addr2_norm),
                       similarity(compact_alnum(addr1_norm),
                                  compact_alnum(addr2_norm)))
        city1_norm = normalize_city(city1)
        city2_norm = normalize_city(city2)
        if city1_norm and city2_norm:
            city_sim = max(similarity(city1_norm, city2_norm),
                          similarity(compact_alnum(city1_norm),
                                     compact_alnum(city2_norm)))
        else:
            city_sim = 0.90
        zip_match = (zip1 and zip2 and zip1 == zip2)
        score = addr_sim * (0.6 + 0.4 * city_sim)
        if zip_match and score >= 0.85:
            score = min(1.0, score + 0.05)

        # Penalty: embedded numbers (2+ digits) that don't overlap (ignore dates)
        nums1 = extract_addr_numbers(addr1_norm)
        nums2 = extract_addr_numbers(addr2_norm)
        nums_mismatch = (nums1 and nums2 and not nums1 & nums2)
        if nums_mismatch:
            score = max(0.0, score - 0.10)

        # Penalty: both have zip codes but they don't match
        zip_penalty = (zip1 and zip2 and zip1 != zip2)
        if zip_penalty:
            score = max(0.0, score - 0.20)

        # Bonus: all embedded numbers match AND zip codes match
        nums_all_match = (nums1 and nums2 and nums1 == nums2)
        if nums_all_match and zip_match:
            score = min(1.0, score + 0.50)

        same = (addr_sim >= 0.90 and city_sim >= 0.85 and not nums_mismatch) or \
               (nums_all_match and zip_match and city_sim >= 0.85)
        if match_detail is not None:
            parts = [f"non_standard: addr_sim={addr_sim:.2f}; city_sim={city_sim:.2f}; zip_match={bool(zip_match)}"]
            if nums_mismatch:
                parts.append("nums_mismatch: -0.10")
            if zip_penalty:
                parts.append("zip_penalty: -0.20")
            if nums_all_match and zip_match:
                parts.append("nums_all_match+zip: +0.50")
            parts.append(f"score={score:.3f}")
            match_detail.extend(parts)
        return _finalize({
            "same_address": same, "score": score,
            "reason": f"NON_STANDARD addr_sim={addr_sim:.2f} "
                      f"city_sim={city_sim:.2f} zip_match={zip_match}"
                      f" nums_mismatch={nums_mismatch}"
                      f" zip_penalty={zip_penalty}"
        })

    if not num1 or not num2:
        if match_detail is not None:
            match_detail.append(f"house_num: {num1 or 'None'} vs {num2 or 'None'} MISSING")
        return _finalize({"same_address": False, "score": 0.0,
                "reason": "MISSING_HOUSE_NUMBER"})
    if num1 != num2:
        if match_detail is not None:
            match_detail.append(f"house_num: {num1}!={num2}")
        return _finalize({"same_address": False, "score": 0.0,
                "reason": f"HOUSE_NUM_MISMATCH {num1}!={num2}"})

    core1 = street_core_for_match(addr1_norm)
    core2 = street_core_for_match(addr2_norm)
    # Fallback: if both cores are empty (street name IS a direction/type,
    # e.g. "N HWY N", "E ST", "S CIR"), compare full post-house-number text
    if not core1 and not core2:
        core1 = re.sub(r"^\s*\d+\s+", "", addr1_norm)
        core2 = re.sub(r"^\s*\d+\s+", "", addr2_norm)
    street_sim = max(similarity(core1, core2),
                     similarity(compact_alnum(core1), compact_alnum(core2)))

    city1_norm = normalize_city(city1)
    city2_norm = normalize_city(city2)
    if city1_norm and city2_norm:
        city_sim = max(similarity(city1_norm, city2_norm),
                       similarity(compact_alnum(city1_norm),
                                  compact_alnum(city2_norm)))
    else:
        city_sim = 0.90

    zip_match = (zip1 and zip2 and zip1 == zip2)
    zip_ok = True
    if zip1 and zip2 and zip1 != zip2:
        zip_ok = city_sim >= 0.92

    same = ((street_sim >= 0.90) and zip_ok and
            (city_sim >= 0.85 or not city1_norm or not city2_norm))
    score = street_sim * (0.6 + 0.4 * city_sim)
    if zip_match and score >= 0.85:
        score = min(1.0, score + 0.05)
    if not zip_ok:
        score *= 0.75

    if match_detail is not None:
        parts = [f"street: '{core1}' vs '{core2}' sim={street_sim:.2f}; city_sim={city_sim:.2f}; zip_ok={zip_ok}"]
        if zip_match and score >= 0.85:
            parts.append("zip_bonus: +0.05")
        if not zip_ok:
            parts.append("zip_penalty: x0.75")
        parts.append(f"score={score:.3f}")
        match_detail.extend(parts)

    return _finalize({
        "same_address": same,
        "score": score,
        "reason": f"STREET num={num1} street_sim={street_sim:.2f} "
                  f"city_sim={city_sim:.2f} zip_ok={zip_ok}"
    })


def _build_addr_detail(a1_detail, a2_detail, match_detail):
    """Build the address detail strings from collected lists."""
    norm_parts = []
    if a1_detail:
        norm_parts.append(f"source: {'; '.join(a1_detail)}")
    if a2_detail:
        norm_parts.append(f"dec: {'; '.join(a2_detail)}")
    return {
        'addr_norm_detail': ' | '.join(norm_parts) if norm_parts else '',
        'addr_match_detail': '; '.join(match_detail) if match_detail else '',
    }


# ===== SSN cleaning =====

BAD_SSNS = {'000000000', '111111111', '222222222', '333333333',
            '444444444', '555555555', '666666666', '777777777',
            '888888888', '999999999'}


def round5(x: float) -> float:
    """Round to nearest 5 (e.g. 3->5, 12->10, 34->35, 89->90)."""
    return round(x / 5) * 5


def mask_ssn(ssn: str) -> str:
    """Mask SSN to show only last 4 digits: 123456789 -> *****6789."""
    digits = re.sub(r'[^0-9]', '', str(ssn)) if ssn else ''
    if len(digits) <= 4:
        return digits
    return '*' * (len(digits) - 4) + digits[-4:]


def clean_ssn(ssn_raw) -> str:
    if not ssn_raw:
        return ""
    digits = re.sub(r'[^0-9]', '', str(ssn_raw))
    if len(digits) < 4:
        return ""
    if digits.startswith('000') or digits.startswith('666'):
        return ""
    if digits in BAD_SSNS:
        return ""
    return digits


# ===== API Override =====

def api_override_name_scores(pairs):
    """Send ambiguous name pairs to Claude API for semantic evaluation.

    Args:
        pairs: list of (index_into_results, source_name, dec_name)
    Returns:
        dict mapping result index -> new score (0-100)
    """
    try:
        import anthropic
    except ImportError:
        print('  anthropic package not installed, skipping API override')
        return {}

    client = anthropic.Anthropic()
    overrides = {}

    for batch_start in range(0, len(pairs), API_BATCH_SIZE):
        batch = pairs[batch_start:batch_start + API_BATCH_SIZE]
        lines = []
        for i, (idx, name1, name2) in enumerate(batch, 1):
            lines.append(f'{i}. "{name1}" vs "{name2}"')

        prompt = (
            "You are evaluating whether business/person name pairs refer to the "
            "same entity. These pairs already share the same SSN, so they are "
            "likely the same person or company.\n\n"
            "For each pair, score 0-100:\n"
            "  100 = definitely same entity (exact or trivial variation)\n"
            "  85-95 = very likely same (abbreviations, nicknames, minor differences)\n"
            "  60-80 = probably same but significant differences\n"
            "  30-55 = uncertain, could be different entities sharing SSN\n"
            "  0-25 = likely different entities\n\n"
            "Name pairs:\n" + "\n".join(lines) + "\n\n"
            "Respond with ONLY a JSON object: {\"scores\": [score1, score2, ...]}"
        )

        try:
            response = client.messages.create(
                model=API_MODEL,
                max_tokens=256,
                temperature=0.0,
                system=(
                    "You are a business entity name matcher for oil and gas "
                    "industry records. You compare name pairs that share the "
                    "same SSN/TIN to determine if they refer to the same "
                    "entity. Be conservative — different companies that happen "
                    "to share a word like 'SOFTWARE' or 'ENERGY' are NOT the "
                    "same entity. Only score high when the names are clearly "
                    "variations of the same entity (abbreviations, nicknames, "
                    "DBA names, trust vs individual, etc)."
                ),
                messages=[{"role": "user", "content": prompt}],
            )
            text = response.content[0].text.strip()
            # Parse JSON from response (handle markdown code blocks)
            if text.startswith('```'):
                text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
            data = json.loads(text)
            scores = data['scores']

            for i, (idx, _, _) in enumerate(batch):
                if i < len(scores):
                    overrides[idx] = round5(max(0, min(100, scores[i])))
        except Exception as e:
            print(f'  API batch {batch_start // API_BATCH_SIZE + 1} failed: {e}')
            err_str = str(e).lower()
            if 'authentication_error' in err_str or '401' in err_str \
               or 'credit balance' in err_str or 'billing' in err_str:
                print('  Fatal API error — skipping remaining batches.')
                print('  Check your API key and account billing at console.anthropic.com.')
                break
            continue

    return overrides


# ===== Google Address Validation API =====

_google_addr_cache = {}  # keyed on (address, city, state, zip) → dict
_google_api_calls = 0




def google_validate_address(address, city, state, zip_code):
    """Validate/standardize a single address via Google Address Validation API.

    Returns dict with keys: standardized, verdict, status.
    Uses in-memory cache first.
    """
    global _google_api_calls

    key = (address, city, state, zip_code)
    if key in _google_addr_cache:
        return _google_addr_cache[key]

    try:
        import requests as req
    except ImportError:
        return {'standardized': None, 'verdict': None, 'status': 'failed'}

    url = f'https://addressvalidation.googleapis.com/v1:validateAddress?key={GOOGLE_API_KEY}'
    body = {
        'address': {
            'regionCode': 'US',
            'addressLines': [address],
            'locality': city,
            'administrativeArea': state,
            'postalCode': zip_code,
        }
    }

    try:
        resp = req.post(url, json=body, timeout=10)
        _google_api_calls += 1

        if resp.status_code in (403, 429):
            return {'standardized': None, 'verdict': None, 'status': 'quota_error'}

        if resp.status_code != 200:
            result = {'standardized': None, 'verdict': None, 'status': 'failed'}
            _google_addr_cache[key] = result
            return result

        data = resp.json()
        api_result = data.get('result', {})
        verdict = api_result.get('verdict')
        postal = api_result.get('address', {}).get('postalAddress', {})

        # Build address from components to exclude point_of_interest (name)
        # that Google bakes into addressLines
        components = api_result.get('address', {}).get('addressComponents', [])
        comp_map = {}
        for comp in components:
            ctype = comp.get('componentType', '')
            ctext = comp.get('componentName', {}).get('text', '')
            if ctext:
                comp_map[ctype] = ctext

        # Reconstruct street address from components (skip point_of_interest)
        parts = []
        if 'street_number' in comp_map:
            parts.append(comp_map['street_number'])
        if 'route' in comp_map:
            parts.append(comp_map['route'])
        if 'subpremise' in comp_map:
            parts.append(comp_map['subpremise'])
        if parts:
            std_addr = ' '.join(parts)
        else:
            # Fallback to addressLines if no components
            std_lines = postal.get('addressLines', [])
            std_addr = std_lines[0] if std_lines else address

        std_city = postal.get('locality', city)
        std_state = postal.get('administrativeArea', state)
        std_zip = postal.get('postalCode', zip_code)
        if std_zip and len(std_zip) > 5:
            std_zip = std_zip[:5]

        standardized = (std_addr.upper(), std_city.upper(),
                        std_state.upper(), std_zip)
        result = {
            'standardized': standardized,
            'verdict': verdict,
            'status': 'success',
        }
        _google_addr_cache[key] = result
        return result

    except Exception:
        result = {'standardized': None, 'verdict': None, 'status': 'failed'}
        _google_addr_cache[key] = result
        return result




def google_override_address_scores(pairs):
    """Send ambiguous address pairs to Google Address Validation API.

    For each pair, validates both Source and DEC addresses, then re-runs
    address_compare() on the standardized forms.

    Args:
        pairs: list of (index, source_addr, source_city, source_state, source_zip,
                        source_id, source_addrseq,
                        dec_addr, dec_city, dec_state, dec_zip,
                        dec_hdrcode, dec_addrsubcode, original_score)
    Returns:
        dict mapping result index -> override info dict with keys:
        new_score, same_address, source_std, source_verdict,
        dec_std, dec_verdict, score_changed
    """
    global _google_api_calls
    _google_api_calls = 0
    overrides = {}
    cache_hits_before = len(_google_addr_cache)
    quota_error = False

    for count, pair in enumerate(pairs, 1):
        if quota_error:
            break

        (idx, c_addr, c_city, c_state, c_zip, c_id, c_addrseq,
         d_addr, d_city, d_state, d_zip, d_hdrcode, d_addrsubcode,
         orig_score) = pair

        # Validate Source address
        c_result = google_validate_address(
            c_addr, c_city, c_state, c_zip)
        if c_result['status'] == 'quota_error':
            print('  Google API quota/auth error — stopping remaining calls.')
            print('  Check your API key and billing at console.cloud.google.com.')
            quota_error = True
            break

        # Validate DEC address
        d_result = google_validate_address(
            d_addr, d_city, d_state, d_zip)
        if d_result['status'] == 'quota_error':
            print('  Google API quota/auth error — stopping remaining calls.')
            quota_error = True
            break

        # Build override entry for this pair (even if score doesn't change)
        entry = {
            'source_std': c_result['standardized'],
            'source_verdict': c_result['verdict'],
            'dec_std': d_result['standardized'],
            'dec_verdict': d_result['verdict'],
        }

        # Re-compare if both sides succeeded
        if c_result['standardized'] and d_result['standardized']:
            c_std = c_result['standardized']
            d_std = d_result['standardized']
            new_compare = address_compare(c_std[0], c_std[1], c_std[3],
                                          d_std[0], d_std[1], d_std[3])
            new_score = round5(new_compare['score'] * 100)
            entry['new_score'] = new_score
            entry['same_address'] = new_compare['same_address']
            entry['score_changed'] = abs(new_score - orig_score) >= 10
        else:
            entry['new_score'] = orig_score
            entry['same_address'] = None
            entry['score_changed'] = False

        overrides[idx] = entry

        if count % 100 == 0:
            cache_hits = len(_google_addr_cache) - cache_hits_before
            print(f'  Progress: {count}/{len(pairs)} pairs '
                  f'({_google_api_calls} API calls, ~{cache_hits} cache hits)')

    cache_hits = len(_google_addr_cache) - cache_hits_before
    print(f'  Google API: {_google_api_calls} calls, '
          f'{len(_google_addr_cache)} cached addresses')

    return overrides


# ===== Main =====

def _load_dec_from_snowflake(sf_conn):
    """Load DEC BA records directly from Enertia views in Snowflake."""
    cur = sf_conn.cursor()
    cur.execute("""
        SELECT b.SSN, a.HDRCODE, a.HDRNAME, a.ADDRADDRESS, a.ADDRCITY,
               a.ADDRSTATE, a.ADDRZIPCODE, a.ADDRCONTACT, a.ADDRSUBCODE
        FROM DGO.ENERTIA.VW_BUSINESS_ASSOCIATE_ADDRESS a
        JOIN DGO.ENERTIA.VW_BUSINESS_ASSOCIATE b
          ON a.HDRCODE = b.BUS_ASSOC_CODE
        WHERE b.ATT_TYPE = 'TaxInfo' AND a.HDRTYPECODE = 'BusAssoc'
    """)
    dec_by_ssn = {}
    dec_total = 0
    for row in cur:
        dec_total += 1
        # Clean embedded newlines (same as load_dec_master.py)
        vals = []
        for v in row:
            if v and isinstance(v, str):
                v = v.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
                v = re.sub(r'\s+', ' ', v).strip()
            vals.append(v or '')
        ssn_clean = clean_ssn(vals[0])
        if ssn_clean:
            if ssn_clean not in dec_by_ssn:
                dec_by_ssn[ssn_clean] = []
            dec_by_ssn[ssn_clean].append({
                'ssn': vals[0], 'hdrcode': vals[1], 'hdrname': vals[2],
                'addraddress': vals[3], 'addrcity': vals[4],
                'addrstate': vals[5], 'addrzipcode': vals[6],
                'addrcontact': vals[7], 'addrsubcode': vals[8],
            })
    cur.close()
    return dec_by_ssn, dec_total


def _write_results_to_snowflake(results_df, sf_conn, run_id):
    """Write match results to Snowflake IMPORT_MERGE_MATCHES table."""
    from snowflake.connector.pandas_tools import write_pandas

    cur = sf_conn.cursor()
    # Create archive table if it doesn't exist (same schema minus autoincrement)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS BA_PROCESS.IMPORT_MERGE_MATCHES_ARCHIVE
        LIKE BA_PROCESS.IMPORT_MERGE_MATCHES
    """)
    # Ensure NAMEADDRSCORE column exists (added after initial table creation)
    for tbl in ('IMPORT_MERGE_MATCHES', 'IMPORT_MERGE_MATCHES_ARCHIVE'):
        try:
            cur.execute(f"ALTER TABLE BA_PROCESS.{tbl} ADD COLUMN NAMEADDRSCORE FLOAT")
        except Exception:
            pass  # column already exists
    # Ensure detail/audit columns exist
    for col in ('NAME_NORMAL_DETAIL', 'ADDRESS_NORMAL_DETAIL',
                'NAME_MATCH_DETAIL', 'ADDR_MATCH_DETAIL'):
        for tbl in ('IMPORT_MERGE_MATCHES', 'IMPORT_MERGE_MATCHES_ARCHIVE'):
            try:
                cur.execute(f"ALTER TABLE BA_PROCESS.{tbl} ADD COLUMN {col} VARCHAR")
            except Exception:
                pass  # column already exists
    # Migrate CANVAS_* columns to SOURCE_* (idempotent — fails silently if already renamed)
    for old_col, new_col in (
        ('CANVAS_NAME', 'SOURCE_NAME'), ('CANVAS_ADDRESS', 'SOURCE_ADDRESS'),
        ('CANVAS_CITY', 'SOURCE_CITY'), ('CANVAS_STATE', 'SOURCE_STATE'),
        ('CANVAS_ZIP', 'SOURCE_ZIP'), ('CANVAS_ADDRSEQ', 'SOURCE_ADDRSEQ'),
        ('CANVAS_ID', 'SOURCE_ID'), ('CANVAS_SSN', 'SOURCE_SSN'),
        ('CANVAS_LOOKUP_VERDICT', 'SOURCE_LOOKUP_VERDICT'),
    ):
        for tbl in ('IMPORT_MERGE_MATCHES', 'IMPORT_MERGE_MATCHES_ARCHIVE'):
            try:
                cur.execute(f"ALTER TABLE BA_PROCESS.{tbl} RENAME COLUMN {old_col} TO {new_col}")
            except Exception:
                pass  # column already renamed or doesn't exist
    # Archive current results before truncating
    cur.execute("""
        INSERT INTO BA_PROCESS.IMPORT_MERGE_MATCHES_ARCHIVE
        SELECT * FROM BA_PROCESS.IMPORT_MERGE_MATCHES
    """)
    cur.execute("TRUNCATE TABLE BA_PROCESS.IMPORT_MERGE_MATCHES")
    sf_conn.commit()

    # Prepare DataFrame for Snowflake (uppercase columns, add RUN_ID)
    sf_df = results_df.copy()
    sf_df.columns = [c.upper() for c in sf_df.columns]
    sf_df['RUN_ID'] = run_id

    success, nchunks, nrows, _ = write_pandas(
        sf_conn, sf_df, 'IMPORT_MERGE_MATCHES',
        schema='BA_PROCESS', database='DGO_MA',
        auto_create_table=False)
    cur.close()
    return nrows


def _classify(name_score, addr_score, bucket_rules):
    """Classify record into bucket based on score ranges. Returns bucket name."""
    for bucket in bucket_rules:
        if (bucket['name_min'] <= name_score <= bucket['name_max'] and
                bucket['addr_min'] <= addr_score <= bucket['addr_max']):
            return bucket['name']
    return 'NEEDS REVIEW'


def main():
    global USE_API_OVERRIDE, API_MODEL, API_SCORE_MIN, API_SCORE_MAX, API_BATCH_SIZE
    global USE_GOOGLE_ADDRESS_API, GOOGLE_ADDR_SCORE_MIN, GOOGLE_ADDR_SCORE_MAX
    global NICKNAMES, BUSINESS_SUFFIXES, BUSINESS_DESCRIPTORS, INDUSTRY_TERMS
    global NAME_SUFFIXES, STATE_ABBREVS, BAD_SSNS

    print('=' * 80)
    print('SOURCE-TO-DEC BA MATCHING')
    print('=' * 80)

    # 0. Snowflake connection (optional)
    sf_conn = None
    if SNOWFLAKE_ENABLED:
        try:
            print('\nConnecting to Snowflake (check browser for SSO)...')
            sf_conn = get_snowflake_connection()
            cur = sf_conn.cursor()
            cur.execute("SELECT CURRENT_DATABASE(), CURRENT_SCHEMA()")
            db, schema = cur.fetchone()
            cur.close()
            print(f'  Connected: {db}.{schema}')
        except Exception as e:
            print(f'  Snowflake unavailable: {e}')
            raise
    else:
        print('\nSnowflake disabled (SNOWFLAKE_ENABLED=false)')

    # 0b. Load config and lookups (Snowflake or hardcoded defaults)
    print('\nLoading configuration...')
    cfg = load_config(sf_conn)
    lookups = load_lookups(sf_conn)

    # Apply config to module globals
    USE_API_OVERRIDE = cfg.get('USE_API_OVERRIDE', USE_API_OVERRIDE)
    API_MODEL = cfg.get('API_MODEL', API_MODEL)
    API_SCORE_MIN = cfg.get('API_SCORE_MIN', API_SCORE_MIN)
    API_SCORE_MAX = cfg.get('API_SCORE_MAX', API_SCORE_MAX)
    API_BATCH_SIZE = int(cfg.get('API_BATCH_SIZE', API_BATCH_SIZE))
    GOOGLE_ADDR_SCORE_MIN = cfg.get('GOOGLE_ADDR_SCORE_MIN', GOOGLE_ADDR_SCORE_MIN)
    GOOGLE_ADDR_SCORE_MAX = cfg.get('GOOGLE_ADDR_SCORE_MAX', GOOGLE_ADDR_SCORE_MAX)

    # Build bucket classification rules from config
    BUCKET_RULES = [
        {
            'name': 'NEW BA AND NEW ADDRESS',
            'stat': 'new_ba_new_addr',
            'name_min': cfg.get('NEW_BA_NEW_ADDR_MIN_NAME_SCORE', 0),
            'name_max': cfg.get('NEW_BA_NEW_ADDR_MAX_NAME_SCORE', 0),
            'addr_min': cfg.get('NEW_BA_NEW_ADDR_MIN_ADDR_SCORE', 0),
            'addr_max': cfg.get('NEW_BA_NEW_ADDR_MAX_ADDR_SCORE', 0),
        },
        {
            'name': 'EXISTING BA ADD NEW ADDRESS',
            'stat': 'existing_ba_new_addr',
            'name_min': cfg.get('EXISTING_BA_NEW_ADDR_MIN_NAME_SCORE', 100),
            'name_max': cfg.get('EXISTING_BA_NEW_ADDR_MAX_NAME_SCORE', 100),
            'addr_min': cfg.get('EXISTING_BA_NEW_ADDR_MIN_ADDR_SCORE', 0),
            'addr_max': cfg.get('EXISTING_BA_NEW_ADDR_MAX_ADDR_SCORE', 0),
        },
        {
            'name': 'EXISTING BA AND EXISTING ADDRESS',
            'stat': 'existing_ba_existing_addr',
            'name_min': cfg.get('EXISTING_BA_EXISTING_ADDR_MIN_NAME_SCORE', 100),
            'name_max': cfg.get('EXISTING_BA_EXISTING_ADDR_MAX_NAME_SCORE', 100),
            'addr_min': cfg.get('EXISTING_BA_EXISTING_ADDR_MIN_ADDR_SCORE', 100),
            'addr_max': cfg.get('EXISTING_BA_EXISTING_ADDR_MAX_ADDR_SCORE', 100),
        },
    ]

    # Apply lookups to module globals
    if 'NICKNAME' in lookups:
        NICKNAMES = lookups['NICKNAME']
    if 'BUSINESS_SUFFIX' in lookups:
        BUSINESS_SUFFIXES = lookups['BUSINESS_SUFFIX']
    if 'BUSINESS_DESCRIPTOR' in lookups:
        BUSINESS_DESCRIPTORS = lookups['BUSINESS_DESCRIPTOR']
    if 'INDUSTRY_TERM' in lookups:
        INDUSTRY_TERMS = lookups['INDUSTRY_TERM']
    if 'NAME_SUFFIX' in lookups:
        NAME_SUFFIXES = lookups['NAME_SUFFIX']
    if 'STATE_ABBREV' in lookups:
        STATE_ABBREVS = lookups['STATE_ABBREV']
    if 'BAD_SSN' in lookups:
        BAD_SSNS = lookups['BAD_SSN']
    if 'TRUST_KEYWORD' in lookups:
        TRUST_KEYWORDS = lookups['TRUST_KEYWORD']
    else:
        TRUST_KEYWORDS = {'TRUST', 'TRUSTEE', 'TRUSTEES', 'TRUSTEESHIP', 'TTEE',
                          'LIVING TRUST', 'FAMILY TRUST', 'REVOCABLE TRUST',
                          'IRREVOCABLE TRUST', 'TESTAMENTARY', 'ESTATE OF', 'DECEDENT'}

    # Generate run_id for this execution
    run_id = time.strftime('%Y%m%d_%H%M%S')

    # 1. Read Source data
    if sf_conn:
        print('\nReading Source data from Snowflake IMPORT_BA_DATA...')
        source_df = pd.read_sql('SELECT * FROM IMPORT_BA_DATA', sf_conn)
        source_df = source_df.fillna('')
        # Snowflake returns uppercase column names — normalise to match Excel expectations
        source_df.columns = [c.upper() for c in source_df.columns]
        print(f'Loaded {len(source_df):,} Source records from Snowflake')
    else:
        print(f'\nReading Source file: {SOURCE_FILE}')
        source_df = pd.read_excel(SOURCE_FILE, dtype=str)
        source_df = source_df.fillna('')
        print(f'Loaded {len(source_df):,} Source records from Excel')

    # 2. Load DEC records into memory dict keyed by clean SSN
    print(f'\nLoading DEC records from Snowflake Enertia views...')
    dec_by_ssn, dec_total = _load_dec_from_snowflake(sf_conn)

    print(f'Loaded {dec_total:,} DEC records')
    dec_with_ssn = sum(len(v) for v in dec_by_ssn.values())
    print(f'DEC records with valid SSN: {dec_with_ssn:,}')
    print(f'Unique DEC SSNs: {len(dec_by_ssn):,}')


    # 4. Process each Source record
    results = []
    stats = {b['stat']: 0 for b in BUCKET_RULES}
    stats['needs_review'] = 0
    _REC_TO_STAT = {b['name']: b['stat'] for b in BUCKET_RULES}
    _REC_TO_STAT['NEEDS REVIEW'] = 'needs_review'

    # Build trust regex from BA_LOOKUP keywords (longest first so multi-word phrases match)
    _trust_pattern = '|'.join(re.escape(kw) for kw in sorted(TRUST_KEYWORDS, key=len, reverse=True))
    _trust_re = re.compile(r'\b(' + _trust_pattern + r')\b', re.IGNORECASE)

    start = time.time()
    total = len(source_df)

    for idx, row in source_df.iterrows():
        source_ssn_raw = row.get('SSN', '')
        source_ssn = clean_ssn(source_ssn_raw)
        source_name = row.get('ENTITY_LIST_NAME', '') or row.get('HDRNAME', '')
        source_addr = row.get('ADDRADDRESS', '')
        source_city = row.get('ADDRCITY', '')
        source_state = row.get('ADDRSTATE', '')
        source_zip = row.get('ADDRZIPCODE', '')
        source_id = row.get('ID', '')
        source_addrseq = row.get('ADDRSEQ', '')

        # Store digits-only SSN (cleaned of dashes, spaces, special chars)
        source_ssn_digits = re.sub(r'[^0-9]', '', str(source_ssn_raw)) if source_ssn_raw else ''

        is_trust = 1 if (_trust_re.search(str(source_name))
                         or _trust_re.search(str(source_addr))) else 0

        base_record = {
            'source_name': source_name,
            'source_address': source_addr,
            'source_city': source_city,
            'source_state': source_state,
            'source_zip': source_zip,
            'source_addrseq': source_addrseq,
            'source_id': source_id,
            'source_ssn': source_ssn_digits,
            'is_trust': is_trust,
            # Google Address Lookup defaults
            'before_lookup_address': None,
            'before_lookup_city': None,
            'before_lookup_state': None,
            'before_lookup_zip': None,
            'address_looked_up': 0,
            'source_lookup_verdict': None,
            'before_lookup_dec_address': None,
            'before_lookup_dec_city': None,
            'before_lookup_dec_state': None,
            'before_lookup_dec_zip': None,
            'dec_address_looked_up': 0,
            'dec_lookup_verdict': None,
        }

        if not source_ssn:
            reason = 'INVALID SSN' if source_ssn_raw else 'NO SSN'
            bad = _needs_review_override(source_name, source_addr,
                                         source_city, source_state, source_zip)
            if bad:
                no_ssn_rec = 'NEEDS REVIEW'
                reason = bad
            else:
                no_ssn_rec = 'NEW BA AND NEW ADDRESS'
            results.append({
                **base_record,
                'dec_hdrcode': '', 'dec_name': '', 'dec_address': '',
                'dec_city': '', 'dec_state': '', 'dec_zip': '',
                'dec_contact': '', 'dec_addrsubcode': '',
                'ssn_match': 0.0, 'name_score': 0.0, 'address_score': 0.0,
                'address_reason': reason, 'nameaddrscore': None,
                'recommendation': no_ssn_rec, 'dec_match_count': 0,
                'Number_possible_address_matches': 0, 'name_boost_source': None,
                'name_normal_detail': '', 'address_normal_detail': '',
                'name_match_detail': '', 'addr_match_detail': '',
            })
            continue

        # Look up in DEC by cleaned SSN
        dec_matches = dec_by_ssn.get(source_ssn, [])

        if not dec_matches:
            bad = _needs_review_override(source_name, source_addr,
                                         source_city, source_state, source_zip)
            if bad:
                no_dec_rec = 'NEEDS REVIEW'
                no_dec_reason = bad
            else:
                no_dec_rec = 'NEW BA AND NEW ADDRESS'
                no_dec_reason = 'SSN NOT FOUND IN DEC'
            results.append({
                **base_record,
                'dec_hdrcode': '', 'dec_name': '', 'dec_address': '',
                'dec_city': '', 'dec_state': '', 'dec_zip': '',
                'dec_contact': '', 'dec_addrsubcode': '',
                'ssn_match': 0.0, 'name_score': 0.0, 'address_score': 0.0,
                'address_reason': no_dec_reason, 'nameaddrscore': None,
                'recommendation': no_dec_rec,
                'dec_match_count': 0,
                'Number_possible_address_matches': 0, 'name_boost_source': None,
                'name_normal_detail': '', 'address_normal_detail': '',
                'name_match_detail': '', 'addr_match_detail': '',
            })
            continue

        # SSN match found - BA is same (100% BA score)
        # Find best address match among all DEC records with this SSN
        best_addr_score = -1
        best_dec = None
        best_addr_result = None
        best_name_result = None
        best_name_score = 0.0
        best_boost_source = None
        possible_addr_matches = 0

        # Pre-extract names from Source address (once per Source record)
        source_addr_names = extract_names_from_address(source_addr)

        for dec in dec_matches:
            addr_result = address_compare(
                source_addr, source_city, source_zip,
                dec['addraddress'], dec['addrcity'], dec['addrzipcode'],
                collect_detail=True)
            name_result = name_compare(source_name, dec['hdrname'],
                                       collect_detail=True)
            primary_score = name_result['name_score']

            # --- Supplementary name comparisons (boost only) ---
            effective_score = primary_score
            boost_source = None

            if primary_score < 0.95:
                supp_candidates = []
                dec_contact = safe_str(dec.get('addrcontact', ''))
                dec_addr_names = extract_names_from_address(
                    dec.get('addraddress', ''))

                # 1) Source name vs DEC addrcontact
                if dec_contact:
                    s = name_compare(source_name, dec_contact)['name_score']
                    if s > effective_score:
                        supp_candidates.append(('DEC_CONTACT', s))

                # 2) Source addr name vs DEC hdrname
                for aname in source_addr_names:
                    s = name_compare(aname, dec['hdrname'])['name_score']
                    if s > effective_score:
                        supp_candidates.append(('SOURCE_ADDR_NAME', s))

                # 3) Source addr name vs DEC addrcontact
                if dec_contact:
                    for aname in source_addr_names:
                        s = name_compare(aname, dec_contact)['name_score']
                        if s > effective_score:
                            supp_candidates.append(
                                ('SOURCE_ADDR+DEC_CONTACT', s))

                # 4) Source name vs DEC addr name
                for dname in dec_addr_names:
                    s = name_compare(source_name, dname)['name_score']
                    if s > effective_score:
                        supp_candidates.append(('DEC_ADDR_NAME', s))

                # 5) Source addr name vs DEC addr name
                for aname in source_addr_names:
                    for dname in dec_addr_names:
                        s = name_compare(aname, dname)['name_score']
                        if s > effective_score:
                            supp_candidates.append(
                                ('SOURCE_ADDR+DEC_ADDR', s))

                # Pick best supplementary score, cap at 0.95
                for src, score in supp_candidates:
                    capped = min(score, 0.95)
                    if capped > effective_score:
                        effective_score = capped
                        boost_source = src

            # Count DEC records exceeding both thresholds (using rounded scores)
            if round5(effective_score * 100) > 94 and round5(addr_result['score'] * 100) > 84:
                possible_addr_matches += 1

            if addr_result['score'] > best_addr_score:
                best_addr_score = addr_result['score']
                best_dec = dec
                best_addr_result = addr_result
                best_name_result = name_result
                best_name_score = effective_score
                best_boost_source = boost_source

        # Data quality gate — force NEEDS REVIEW for bad/missing data
        bad = _needs_review_override(
            source_name, source_addr, source_city, source_state, source_zip,
            dec_name=best_dec['hdrname'], dec_addr=best_dec['addraddress'],
            dec_city=best_dec['addrcity'], dec_state=best_dec['addrstate'],
            dec_zip=best_dec['addrzipcode'])
        if bad or best_addr_result.get('reason') == 'UNKNOWN_ADDRESS':
            rec = 'NEEDS REVIEW'
        else:
            rec = None  # deferred — classify after API overrides

        # Composite name+address score (holistic identity similarity)
        source_combo = (normalize_name(source_name, trust=is_trust) + ' '
                        + normalize_address(source_addr, trust=is_trust) + ' '
                        + normalize_city(source_city) + ' '
                        + safe_str(source_state).upper().strip() + ' '
                        + normalize_zip(source_zip))
        dec_combo = (normalize_name(best_dec['hdrname']) + ' '
                     + normalize_address(best_dec['addraddress']) + ' '
                     + normalize_city(best_dec['addrcity']) + ' '
                     + safe_str(best_dec['addrstate']).upper().strip() + ' '
                     + normalize_zip(best_dec['addrzipcode']))
        nameaddrscore = round5(jaro_winkler(source_combo, dec_combo) * 100)

        # Build detail strings from best match
        _name_norm_detail = best_name_result.get('name_norm_detail', '') if best_name_result else ''
        _name_match_detail = best_name_result.get('name_match_detail', '') if best_name_result else ''
        _addr_norm_detail = best_addr_result.get('addr_norm_detail', '') if best_addr_result else ''
        _addr_match_detail = best_addr_result.get('addr_match_detail', '') if best_addr_result else ''
        if best_boost_source:
            boost_line = f"BOOST({best_boost_source}): {best_name_score:.3f}"
            _name_match_detail = f"{_name_match_detail}; {boost_line}" if _name_match_detail else boost_line

        results.append({
            **base_record,
            'dec_hdrcode': best_dec['hdrcode'],
            'dec_name': best_dec['hdrname'],
            'dec_address': best_dec['addraddress'],
            'dec_city': best_dec['addrcity'],
            'dec_state': best_dec['addrstate'],
            'dec_zip': best_dec['addrzipcode'],
            'dec_contact': best_dec['addrcontact'],
            'dec_addrsubcode': best_dec.get('addrsubcode', ''),
            'ssn_match': 100.0,
            'name_score': round5(best_name_score * 100),
            'address_score': round5(best_addr_score * 100),
            'address_reason': best_addr_result['reason'],
            'nameaddrscore': nameaddrscore,
            'recommendation': rec,
            'name_boost_source': best_boost_source,
            'dec_match_count': len(dec_matches),
            'Number_possible_address_matches': possible_addr_matches,
            'name_normal_detail': _name_norm_detail,
            'address_normal_detail': _addr_norm_detail,
            'name_match_detail': _name_match_detail,
            'addr_match_detail': _addr_match_detail,
        })
        # Also flag as trust if DEC name or address matches
        if not is_trust and (_trust_re.search(str(best_dec['hdrname']))
                             or _trust_re.search(str(best_dec['addraddress']))):
            results[-1]['is_trust'] = 1

        if (idx + 1) % 5000 == 0:
            elapsed = time.time() - start
            rate = (idx + 1) / elapsed
            remaining = (total - idx - 1) / rate
            print(f'  Processed {idx + 1:,}/{total:,} '
                  f'({elapsed:.1f}s elapsed, ~{remaining:.0f}s remaining)')

    elapsed = time.time() - start
    print(f'\nProcessed {total:,} records in {elapsed:.1f}s')

    # 4b. API override for ambiguous name scores
    if USE_API_OVERRIDE:
        ambiguous = [
            (i, r['source_name'], r['dec_name'])
            for i, r in enumerate(results)
            if r['ssn_match'] == 100.0
            and API_SCORE_MIN <= r['name_score'] <= API_SCORE_MAX
        ]
        if ambiguous:
            batches = (len(ambiguous) + API_BATCH_SIZE - 1) // API_BATCH_SIZE
            print(f'\nAPI override: {len(ambiguous)} ambiguous name pairs '
                  f'({batches} batch{"es" if batches != 1 else ""})...')
            overrides = api_override_name_scores(ambiguous)
            for idx, new_score in overrides.items():
                results[idx]['name_score'] = float(new_score)
            print(f'  Overrode {len(overrides)} scores')

    # 4c. Google Address Validation API override for ambiguous address scores
    if USE_GOOGLE_ADDRESS_API and GOOGLE_API_KEY:
        def _zero_score_nums_match(r):
            """Score=0 but all address numbers and zip match → worth a Google lookup."""
            if r['address_score'] != 0:
                return False
            z1, z2 = normalize_zip(r['source_zip']), normalize_zip(r['dec_zip'])
            if not z1 or not z2 or z1 != z2:
                return False
            n1 = extract_addr_numbers(normalize_address(r['source_address']))
            n2 = extract_addr_numbers(normalize_address(r['dec_address']))
            return bool(n1) and n1 == n2

        ambiguous_addr = [
            (i, r['source_address'], r['source_city'], r['source_state'],
             r['source_zip'], r['source_id'], r['source_addrseq'],
             r['dec_address'], r['dec_city'], r['dec_state'],
             r['dec_zip'], r['dec_hdrcode'], r.get('dec_addrsubcode', ''),
             r['address_score'])
            for i, r in enumerate(results)
            if r['ssn_match'] == 100.0
            and (GOOGLE_ADDR_SCORE_MIN <= r['address_score'] <= GOOGLE_ADDR_SCORE_MAX
                 or _zero_score_nums_match(r))
        ]
        if ambiguous_addr:
            print(f'\nGoogle Address API: {len(ambiguous_addr)} ambiguous address pairs...')
            addr_overrides = google_override_address_scores(ambiguous_addr)
            score_overrides = 0
            for idx, ov in addr_overrides.items():
                r = results[idx]

                # Save originals before overwriting (Source)
                r['before_lookup_address'] = r['source_address']
                r['before_lookup_city'] = r['source_city']
                r['before_lookup_state'] = r['source_state']
                r['before_lookup_zip'] = r['source_zip']
                r['address_looked_up'] = 1
                r['source_lookup_verdict'] = (
                    json.dumps(ov['source_verdict'])
                    if ov['source_verdict'] else None)

                # Save originals before overwriting (DEC)
                r['before_lookup_dec_address'] = r['dec_address']
                r['before_lookup_dec_city'] = r['dec_city']
                r['before_lookup_dec_state'] = r['dec_state']
                r['before_lookup_dec_zip'] = r['dec_zip']
                r['dec_address_looked_up'] = 1
                r['dec_lookup_verdict'] = (
                    json.dumps(ov['dec_verdict'])
                    if ov['dec_verdict'] else None)

                # Replace Source address with standardized
                if ov['source_std']:
                    r['source_address'] = ov['source_std'][0]
                    r['source_city'] = ov['source_std'][1]
                    r['source_state'] = ov['source_std'][2]
                    r['source_zip'] = ov['source_std'][3]

                # Replace DEC address with standardized
                if ov['dec_std']:
                    r['dec_address'] = ov['dec_std'][0]
                    r['dec_city'] = ov['dec_std'][1]
                    r['dec_state'] = ov['dec_std'][2]
                    r['dec_zip'] = ov['dec_std'][3]

                # Update score if meaningfully changed
                if ov['score_changed']:
                    score_overrides += 1
                    r['address_score'] = float(ov['new_score'])

            print(f'  Validated {len(addr_overrides)} pairs, '
                  f'{score_overrides} scores overridden')
    elif USE_GOOGLE_ADDRESS_API and not GOOGLE_API_KEY:
        print('\nGoogle Address API enabled but GOOGLE_API_KEY not set — skipping.')

    # 5. Final classification — single pass with best available scores
    stats = {b['stat']: 0 for b in BUCKET_RULES}
    stats['needs_review'] = 0
    for r in results:
        if r['recommendation'] is None:
            r['recommendation'] = _classify(
                float(r['name_score']), float(r['address_score']), BUCKET_RULES)
        stats[_REC_TO_STAT[r['recommendation']]] += 1

    # 6. Insert results into database
    print(f'\nWriting {len(results):,} results to import_merge_matches table...')
    results_df = pd.DataFrame(results)

    # Ensure column order matches table schema
    col_order = [
        'ssn_match', 'name_score', 'name_boost_source', 'address_score', 'address_reason',
        'name_normal_detail', 'address_normal_detail',
        'name_match_detail', 'addr_match_detail',
        'nameaddrscore', 'recommendation',
        'source_name', 'source_address', 'source_city', 'source_state',
        'source_zip', 'source_addrseq', 'source_id', 'source_ssn',
        # Google lookup - Source
        'before_lookup_address', 'before_lookup_city',
        'before_lookup_state', 'before_lookup_zip',
        'address_looked_up', 'source_lookup_verdict',
        # DEC
        'dec_hdrcode', 'dec_name', 'dec_address', 'dec_city',
        'dec_state', 'dec_zip', 'dec_contact', 'dec_addrsubcode',
        # Google lookup - DEC
        'before_lookup_dec_address', 'before_lookup_dec_city',
        'before_lookup_dec_state', 'before_lookup_dec_zip',
        'dec_address_looked_up', 'dec_lookup_verdict',
        #
        'dec_match_count', 'Number_possible_address_matches',
        'is_trust',
    ]
    results_df = results_df[col_order]
    results_df = results_df.sort_values('recommendation').reset_index(drop=True)

    # 7. Write results to Snowflake
    if sf_conn:
        print('\nWriting results to Snowflake...')
        nrows = _write_results_to_snowflake(results_df, sf_conn, run_id)
        print(f'  Wrote {nrows} rows to Snowflake IMPORT_MERGE_MATCHES')

    # 6. Print summary
    print(f'\n{"=" * 80}')
    print('MATCHING RESULTS SUMMARY')
    print('=' * 80)
    print(f'Total Source records:              {total:,}')
    print(f'')
    print(f'NEW BA AND NEW ADDRESS:            {stats["new_ba_new_addr"]:,}')
    print(f'EXISTING BA ADD NEW ADDRESS:       {stats["existing_ba_new_addr"]:,}')
    print(f'EXISTING BA AND EXISTING ADDRESS:  {stats["existing_ba_existing_addr"]:,}')
    print(f'NEEDS REVIEW:                      {stats["needs_review"]:,}')
    print(f'')
    total_matched = (stats['existing_ba_new_addr']
                     + stats['existing_ba_existing_addr']
                     + stats['needs_review'])
    print(f'Total SSN matched:                 {total_matched:,}')
    print(f'Total new BAs:                     {stats["new_ba_new_addr"]:,}')
    print(f'')
    print('Bucket ranges (name_min-max / addr_min-max):')
    for b in BUCKET_RULES:
        print(f'  {b["name"]}: name {b["name_min"]}-{b["name_max"]}'
              f' / addr {b["addr_min"]}-{b["addr_max"]}')
    print('=' * 80)

    # 8. Export to Excel with color-coded headers
    excel_path = 'output/import_merge_matches.xlsx'
    os.makedirs('output', exist_ok=True)
    print(f'\nExporting to {excel_path}...')
    export_df = results_df.copy()
    # Strip control characters that openpyxl rejects
    for col in export_df.select_dtypes(include='object').columns:
        export_df[col] = export_df[col].apply(
            lambda x: re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str(x)) if pd.notna(x) else x)

    # Mask SSNs in Excel output (show last 4 only)
    if 'source_ssn' in export_df.columns:
        export_df['source_ssn'] = export_df['source_ssn'].apply(mask_ssn)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Matches')
        ws = writer.sheets['Matches']
        ws.freeze_panes = 'A2'

        blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1',
                                fill_type='solid')
        tan_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9',
                               fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
        header_font = Font(bold=True)

        # Identify trust-related rows
        trust_words = re.compile(
            r'\b(TRUST|TRUSTEE|TRUSTEES|TRUSTEESHIP|LIVING TRUST|'
            r'FAMILY TRUST|REVOCABLE TRUST|IRREVOCABLE TRUST|'
            r'TESTAMENTARY|ESTATE OF|DECEDENT)\b', re.IGNORECASE)
        trust_rows = set()
        for i, row_data in export_df.iterrows():
            name_text = f"{row_data.get('source_name', '')} {row_data.get('dec_name', '')}"
            if trust_words.search(name_text):
                trust_rows.add(i + 2)  # +2: 1-indexed header + 1-indexed data

        num_cols = len(export_df.columns)
        for col_idx, col_name in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            if col_name.startswith('source_'):
                fill = blue_fill
            elif col_name.startswith('dec_'):
                fill = tan_fill
            else:
                fill = None
            if fill:
                for row_idx in range(1, len(export_df) + 2):
                    if row_idx not in trust_rows:
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        # Apply light red to entire row for trust-related records
        for row_idx in trust_rows:
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

        print(f'  Highlighted {len(trust_rows):,} trust-related rows in light red')

        # ===== Stats tab =====
        # Use unmasked export_df (already has all match data)
        df = export_df  # alias for brevity

        # Pre-compute trust flag on df for stats
        df['_is_trust'] = df.apply(
            lambda r: bool(trust_words.search(
                f"{r.get('source_name', '')} {r.get('dec_name', '')}")),
            axis=1)

        ssn_matched = df[df['ssn_match'] == 100]
        no_match = df[df['ssn_match'] == 0]

        stats_rows = []

        def add(label, value, section=''):
            stats_rows.append({'Section': section, 'Metric': label,
                               'Value': value})

        def add_blank():
            stats_rows.append({'Section': '', 'Metric': '', 'Value': ''})

        # --- Overall ---
        add('Total Source Records', len(df), 'OVERALL')
        add('Unique Source SSNs (non-empty)',
            df[df['source_ssn'].str.len() > 0]['source_ssn'].nunique())
        add('Records with SSN Match (Existing BA)', len(ssn_matched))
        add('Records with No Match (New BA)', len(no_match))
        add_blank()

        # --- Recommendation Breakdown ---
        add('NEW BA AND NEW ADDRESS',
            len(df[df['recommendation'] == 'NEW BA AND NEW ADDRESS']),
            'BY RECOMMENDATION')
        add('EXISTING BA ADD NEW ADDRESS',
            len(df[df['recommendation'] == 'EXISTING BA ADD NEW ADDRESS']))
        add('EXISTING BA AND EXISTING ADDRESS',
            len(df[df['recommendation'] == 'EXISTING BA AND EXISTING ADDRESS']))
        add('NEEDS REVIEW',
            len(df[df['recommendation'] == 'NEEDS REVIEW']))
        add_blank()

        # --- Bucket Ranges ---
        for b in BUCKET_RULES:
            add(f'{b["name"]}: name {b["name_min"]}-{b["name_max"]}'
                f' / addr {b["addr_min"]}-{b["addr_max"]}',
                len(df[df['recommendation'] == b['name']]),
                'BUCKET RANGES' if b == BUCKET_RULES[0] else '')
        add('NEEDS REVIEW (fallback)',
            len(df[df['recommendation'] == 'NEEDS REVIEW']))
        add_blank()

        # --- Name Boost ---
        boosted = ssn_matched[ssn_matched['name_boost_source'].notna()]
        add('Records with Name Score Boosted', len(boosted),
            'NAME BOOST (from address/contact fields)')
        for src in ['DEC_CONTACT', 'SOURCE_ADDR_NAME',
                     'SOURCE_ADDR+DEC_CONTACT', 'DEC_ADDR_NAME',
                     'SOURCE_ADDR+DEC_ADDR']:
            cnt = len(boosted[boosted['name_boost_source'] == src])
            if cnt > 0:
                add(f'  Boost from {src}', cnt)
        add_blank()

        # --- Trust / Estate ---
        trust_df = df[df['_is_trust']]
        add('Total Trust/Estate Records', len(trust_df),
            'TRUST & ESTATE')
        add('Trusts with SSN Match', len(trust_df[trust_df['ssn_match'] == 100]))
        add('Trusts with No Match', len(trust_df[trust_df['ssn_match'] == 0]))
        add('Pct of All Records that are Trust/Estate',
            f"{len(trust_df) / max(len(df), 1) * 100:.1f}%")
        add_blank()

        # --- Multiple DEC Matches ---
        multi = ssn_matched[ssn_matched['dec_match_count'] > 1]
        add('Records with Multiple DEC Matches', len(multi),
            'DUPLICATE RISK')
        add('Records with Possible Address Matches > 0',
            len(df[df['Number_possible_address_matches'] > 0]))
        add('Avg DEC Records per SSN (when matched)',
            f"{ssn_matched['dec_match_count'].mean():.1f}")
        add('Max DEC Records for Single SSN',
            int(ssn_matched['dec_match_count'].max()))
        add_blank()

        # --- Name Score Distribution (SSN-matched only) ---
        add('Name Score = 100', len(ssn_matched[ssn_matched['name_score'] == 100]),
            'NAME SCORE DISTRIBUTION (SSN MATCHED)')
        add('Name Score 90-95',
            len(ssn_matched[ssn_matched['name_score'].between(90, 95)]))
        add('Name Score 80-85',
            len(ssn_matched[ssn_matched['name_score'].between(80, 85)]))
        add('Name Score 70-75',
            len(ssn_matched[ssn_matched['name_score'].between(70, 75)]))
        add('Name Score < 70',
            len(ssn_matched[ssn_matched['name_score'] < 70]))
        add_blank()

        # --- Top States ---
        add('', '', 'TOP 10 STATES')
        top_states = (df[df['source_state'] != '']
                      .groupby('source_state').size()
                      .sort_values(ascending=False).head(10))
        for state, cnt in top_states.items():
            add(state, cnt)

        # Write stats sheet
        stats_df = pd.DataFrame(stats_rows)
        stats_df.to_excel(writer, index=False, sheet_name='Summary Stats')
        ws2 = writer.sheets['Summary Stats']

        # Format stats sheet
        section_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                   fill_type='solid')
        section_font = Font(bold=True, color='FFFFFF')
        metric_font = Font(bold=False)
        value_font = Font(bold=True)
        header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4',
                                  fill_type='solid')

        # Header row
        for col_idx in range(1, 4):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Format section headers and values
        for row_idx in range(2, len(stats_df) + 2):
            section_val = ws2.cell(row=row_idx, column=1).value
            metric_val = ws2.cell(row=row_idx, column=2).value
            if section_val and not metric_val:
                # Section header row
                for col_idx in range(1, 4):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    cell.fill = section_fill
                    cell.font = section_font
            elif metric_val:
                ws2.cell(row=row_idx, column=2).font = metric_font
                ws2.cell(row=row_idx, column=3).font = value_font

        # Column widths
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 18

        # Clean up temp column
        df.drop(columns=['_is_trust'], inplace=True, errors='ignore')

        print(f'  Created Summary Stats tab')

        # ===== Data Completeness tab =====
        total_rows = len(df)
        completeness_rows = []
        for col in df.columns:
            filled = df[col].apply(
                lambda v: v is not None and str(v).strip() != ''
                and str(v).strip().lower() != 'nan').sum()
            pct = filled / total_rows * 100 if total_rows > 0 else 0
            completeness_rows.append({
                'Column': col,
                'Non-Empty Count': int(filled),
                'Empty Count': total_rows - int(filled),
                'Pct Filled': round(pct, 1),
            })

        comp_df = pd.DataFrame(completeness_rows)
        comp_df.to_excel(writer, index=False, sheet_name='Data Completeness')
        ws3 = writer.sheets['Data Completeness']

        # Format header
        for col_idx in range(1, 5):
            cell = ws3.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Color-code Pct Filled column (col 4)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                                 fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C',
                                  fill_type='solid')
        red_comp_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                                    fill_type='solid')
        for row_idx in range(2, len(comp_df) + 2):
            pct_val = ws3.cell(row=row_idx, column=4).value
            if pct_val is not None:
                if pct_val >= 80:
                    ws3.cell(row=row_idx, column=4).fill = green_fill
                elif pct_val >= 40:
                    ws3.cell(row=row_idx, column=4).fill = yellow_fill
                else:
                    ws3.cell(row=row_idx, column=4).fill = red_comp_fill

        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 12

        print(f'  Created Data Completeness tab')

    print(f'Exported {len(export_df):,} rows to {excel_path}')
    print('=' * 80)

    if sf_conn:
        sf_conn.close()
        print('Snowflake connection closed.')


if __name__ == '__main__':
    main()
